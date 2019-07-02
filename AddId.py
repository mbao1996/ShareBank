#coding=gbk
from openpyxl import load_workbook
from define import *
from lib import *

fn = work_catalog + bank_name
fn_rd = work_catalog + '\list.xlsm'

ShareBank = read_data(fn)
wb = load_workbook(fn_rd,keep_vba=True)
ws = wb['Sheet1']

for row in range(1, ws.max_row + 1):
    if( share_exist(ShareBank, ws.cell(row, 1).value[2:8]) == False ):
        s = Share()
        s.id = ws.cell(row, 1).value[2:8]
        share_add(ShareBank, s)

for i in range( len(ShareBank) ):
    s = ShareBank[i]
    print(i, ': ', s.nmcard(),'\n   ', s.dt,'\n   ', s.rt,'\n   ', s.cp)

save_data(fn, ShareBank)
print('\n finished')
