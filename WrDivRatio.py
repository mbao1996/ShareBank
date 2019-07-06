#coding=gbk
from openpyxl import load_workbook
from define import *
from lib import *

fn = work_catalog + bank_name
fn_wr = work_catalog + '\DivRatio.xlsm'

row_title = 1
row_start = row_title + 1
# -------for sheet 'Sheet1' ---------------
col_code = 1
col_name = col_code + 1
col_price = col_name + 1
col_div_ratio = col_price + 1
col_exp_div_ratio = col_div_ratio + 1
col_adjust = col_exp_div_ratio + 1
col_last_div = col_adjust + 1
col_safe_div = col_last_div + 1
col_safe_div_r = col_safe_div + 1
col_agv_div_r = col_safe_div_r + 1
col_div_up = col_agv_div_r + 1
col_qtr_acc = col_div_up + 1
col_pft_3years_inc = col_qtr_acc + 1
col_current_ratio = col_pft_3years_inc + 1
col_gold_incl = col_current_ratio + 1
col_industry = col_gold_incl + 1
col_flag = col_industry + 1
col_last_dividend = col_flag + 1
col_EPS_ttm = col_last_dividend + 1
col_grade = col_EPS_ttm + 1

ShareBank = read_data(fn)
try:
    wb = load_workbook(fn_wr,keep_vba=True)
except Exception as e:
    print(str(e))
    os._exit(0)
ws = wb['Sheet1']

gd_flag = 'grade'

xl_row = row_start

for i in range( len(ShareBank) ):
    s = ShareBank[i]
    if( my_flag in s.flag or hd_flag in s.flag ):
#        print(i, '[max:', len(ShareBank), ']: ', s.nmcard(), '---flag---:', s.flag)
        ws.cell(xl_row, col_code).value = get_sina_id(s.id)
        ws.cell(xl_row, col_name).value = s.name
        ws.cell(xl_row, col_price).value = s.price
#        ws.cell(xl_row, col_div_ratio).value = s.cp['stk_div_ratio']
#        val = s.rt['avg_div_rate'] * s.rt['EPS_ttm'] / s.price
#        ws.cell(xl_row, col_exp_div_ratio).value = round(val, 4)
        ws.cell(xl_row, col_adjust).value = s.rt['convert_rate']
        ws.cell(xl_row, col_last_div).value = s.rt['last_year_div']
#        ws.cell(xl_row, col_safe_div).value = s.cp['hope_div']
#        ws.cell(xl_row, col_safe_div_r).value = s.cp['safe_div']
        ws.cell(xl_row, col_agv_div_r).value = s.rt['avg_div_rate']
        ws.cell(xl_row, col_div_up).value = s.rt['income_up']
        ws.cell(xl_row, col_qtr_acc).value = s.rt['profit_dedt_acc']
        ws.cell(xl_row, col_pft_3years_inc).value = pft_3_years_increase(s)
        ws.cell(xl_row, col_current_ratio).value = s.dt['current_ratio']
        ws.cell(xl_row, col_gold_incl).value = s.rt['gold_include']
        if( 'industry' in s.dt ):
            ws.cell(xl_row, col_industry).value = s.dt['industry']
        ws.cell(xl_row, col_flag).value = fill_flag(s)
        ws.cell(xl_row, col_last_dividend).value = s.dt['dividend'][1]
        ws.cell(xl_row, col_EPS_ttm).value = s.rt['EPS_ttm']
        if( gd_flag in s.flag ):
            ws.cell(xl_row, col_grade).value = s.flag[gd_flag]
        else: 
            ws.cell(xl_row, col_grade).value = -1

        xl_row += 1
try:
    wb.save(fn_wr)
except Exception as e:
    print(str(e))
    os._exit(0)
print('\n finished')
